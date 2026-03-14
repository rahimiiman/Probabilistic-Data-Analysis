import typing as tp
import pandas as pd
import numpy as np
from scipy import stats as stat
import matplotlib.pyplot as plt
import seaborn as sns
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from sklearn.preprocessing import StandardScaler, MinMaxScaler , PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split , cross_val_score
from sklearn.pipeline import Pipeline


class data_cleaning():
    """
        this class check and report summary of data base as an excel file 
        Also it can be used for missing data handing and normalization
    """
    
    def __init__(self,df):
        self.df=df
        
    def database_summary(self,output_name: str , 
                         missing_data_indicator: str | list[str] | None = None):
    
        
        df = self.df
        df_summaary = df.describe(include='all')                                            # Get summary statistics of the data

        #transpose the summary to make it easier to read
        df_summaary = df_summaary.transpose()

        # adding data types of each column to the summary
        df_summaary['data_type'] = df.dtypes

        # calculate the number of null values in each column and add it to the summary
        df_summaary['num_df.isnull'] = df.isnull().sum()

        # calculate number of np.nan values in each column and add it to the summary
        df_summaary['num_np_nan'] = df.apply(lambda x: np.isnan(x).sum() if x.dtype in ['float64', 'int64'] else 0)

        # number of blancks in each column and add it to the summary
        df_summaary['num_blanks'] = df.apply(lambda x: (x == '').sum() if x.dtype == 'object' else 0)

        if missing_data_indicator is not None:
            for item in missing_data_indicator:
                col_name='num_'+str(item)
                df_summaary[col_name]=df.apply(lambda x: (x == item).sum() if x.dtype == 'object' else 0)

        # write the summary to a new CSV file
        df_summaary.to_excel(output_name)
        
        return None
    
    def unify_missing_indicator(self,missing_indicators : list[str] = None):
        
        for item in missing_indicators:
            self.df=self.df.replace(item,np.nan)
    
    def missing_data_handling(self,column_list :list[str]
                              ,missing_data_indicator : str
                              ,action: str ='drop'):
        """"
        column_list : list of columns we want to handle missing data
        missing_data_indicator : for example np.nan
        action : 
                - drop : drop the row
                - mean : replace by mean value of column
                - max : replace with the most frequent value
        """
        
        if isinstance(column_list, str):
            column_list = [column_list]
        
        # Must be list now
        if not isinstance(column_list, list):
            raise TypeError(f"column_list must be str or list[str], got {type(column_list)}")
            return None

        if action== 'drop':
            self.df= self.df.dropna(subset=column_list,axis=0)
            self.df=self.df.reset_index(drop=True)
        
        elif action== 'mean':
            for col in column_list:
                avg_col=self.df[col].astype('float').mean()
                self.df[col]= self.df[col].replace(missing_data_indicator,avg_col)
                
        elif action== 'max':
            for col in column_list:
                most_frequent = self.df[col].mode()[0]
                self.df[col]= self.df[col].replace(missing_data_indicator,most_frequent)
    
    def data_format_handling(self , column_list : list[str],
                             dtype : str ='float64'):
        
        """Examples of valid dtype values:
        'float64', 'float32', 'int64', 'int32', 'category',
        'string', 'bool', 'datetime64[ns]', pd.CategoricalDtype(...)
        
        'float' also refer to 'float64'
        """
        
        if isinstance(column_list, str):
            column_list = [column_list]
        
        # Must be list now
        if not isinstance(column_list, list):
            raise TypeError(f"column_list must be str or list[str], got {type(column_list)}")
            return None
        
        
        for col in column_list:
            self.df[col]=self.df[col].astype(dtype)
        
    def normalization(self,column_list: list[str],
                      action : str ='scaling'):
        """
        used to normalize data
        action can be : 
                        - scaling :: current_value / max
                        - min_max :: (current_value - min)/ (max - min)
                        - zscore  :: (current_value - mu)/sigma
        """
        
        if action == 'scaling':
            for col in column_list:
                max_col = self.df[col].astype('float').max()
                self.df[col] = self.df[col].astype('float') / max_col
                
        elif action == 'min_max':
            for col in column_list:
                min_col = self.df[col].astype('float').min()
                max_col = self.df[col].astype('float').max()
                self.df[col] = (self.df[col].astype('float') - min_col) / (max_col - min_col)
                
        elif action == 'zscore':
            for col in column_list:
                mean_col = self.df[col].astype('float').mean()
                std_col = self.df[col].astype('float').std()
                self.df[col] = (self.df[col].astype('float') - mean_col) / std_col

    def categorical_handling(self,
                                cat_columns: list[str] | None = None,
                                prefix_sep: str = "_",
                                dtype='bool'               # bool True and False or 'int' 0 and 1 used in one-hot mapping
                            ):
        """
        One-hot encodes specified categorical columns and appends the dummy columns
        to the right side of the dataframe with clear naming: original_col_value

        Parameters:
        -----------
        cat_columns : list[str] or None
            List of categorical column names to encode.

        prefix_sep : str, default "_"
            Separator between original column name and category value

        """
        # We'll collect all dummy dataframes here
        dummy_dfs = []

        for col in cat_columns:
            if col not in self.df.columns:
                print(f"Warning: column '{col}' not in dataframe → skipped")
                continue

            # Create dummies for this column
            dummies = pd.get_dummies(
                self.df[col],
                prefix=col,
                prefix_sep=prefix_sep,
                dtype=dtype   # or bool, depending on preference
            )

            dummy_dfs.append(dummies)


        # Combine all dummy columns into one dataframe
        all_dummies = pd.concat(dummy_dfs, axis=1)

        # Append dummies to the right
        self.df = pd.concat([self.df, all_dummies], axis=1)

class DescriptiveStatistics:

    def __init__(self, df):
        self.df = df

    def _save_figures_to_one_sheet(self, plot_func, items, excel_filename, sheet_name, figsize=(6.5, 5.5), vertical_spacing=25):
        """Each plot is its own figure but all go into one Excel sheet stacked vertically."""
        temp_images = []
        for i, item in enumerate(items):
            fig, ax = plt.subplots(figsize=figsize)
            plot_func(item, ax)
            plt.tight_layout()
            temp_img = f"temp_plot_{i}.png"
            fig.savefig(temp_img, bbox_inches="tight", dpi=120)
            plt.close(fig)
            temp_images.append(temp_img)

        # Prepare Excel
        if os.path.exists(excel_filename):
            try:
                book = load_workbook(excel_filename)
                if sheet_name in book.sheetnames:
                    book.remove(book[sheet_name])
            except Exception:
                if os.path.exists(excel_filename):
                    os.remove(excel_filename)
                book = Workbook()
        else:
            book = Workbook()

        if "Sheet" in book.sheetnames and len(book.sheetnames) == 1:
            book.remove(book["Sheet"])

        ws = book.create_sheet(sheet_name)

        # Add images vertically
        current_row = 2
        for temp_img in temp_images:
            img = XLImage(temp_img)
            ws.add_image(img, f"B{current_row}")
            current_row += vertical_spacing

        book.save(excel_filename)
        book.close()

        # Now it's safe to delete temp images
        for temp_img in temp_images:
            if os.path.exists(temp_img):
                os.remove(temp_img)

        print(f"Saved {len(items)} plots → sheet '{sheet_name}'")
        

    def boxplot(self, x_list : list[list[str]] | str ,
                y_list : list[str],
                excel_filename : str ="Descriptive_snapshots.xlsx"):
        """
        x_list : is a list of list of column names like [[x1,x2],[x1,x3,x4],[x2]]
        y_list : is a list of column names [y1,y2,y1]
        both must have same length 
        the data of y1 is grouped for all combination of (x1,x2) and then plot using groupby method of pandas
        
        """

        if isinstance(x_list, str):
            x_list = [[x_list]]
        if isinstance(y_list, str):
            y_list = [y_list]

        if len(x_list) != len(y_list):
            raise ValueError("x_list and y_list must have same length")
        
        x_list_check = isinstance(x_list, list) and all(
                        isinstance(inner, list) and all(isinstance(s, str) for s in inner)
                        for inner in x_list
                        )
        
        if not x_list_check:
            raise ValueError("x_list must be type of list[list[str]]")
        if not isinstance(y_list,list):
            raise ValueError("y_list must be type list")
        
        def plot_one(args, ax):
            x_cols, y_col = args
            self.df.boxplot(column=y_col, by=x_cols, ax=ax)
            ax.set_title(f"{y_col} by {' + '.join(x_cols)}")
            ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right')
            ax.tick_params(axis='x', labelsize=8)

        pairs = list(zip(x_list, y_list))
        self._save_figures_to_one_sheet(plot_one, pairs, excel_filename, sheet_name="Boxplots")

    def scatter_with_regression(self, x_list : list[str], 
                                y_list : list[str],
                                excel_filename : str="Descriptive_snapshots.xlsx"):
        if len(x_list) != len(y_list):
            raise ValueError("x_list and y_list must have same length")
        if not isinstance(x_list,list):
            raise ValueError("x_list must be type list")
        if not isinstance(y_list,list):
            raise ValueError("y_list must be type list")

        def plot_one(args, ax):
            x_col, y_col = args
            sns.regplot(
                data=self.df,
                x=x_col,
                y=y_col,
                ax=ax,
                scatter_kws={'s': 30, 'alpha': 0.7},
                line_kws={'color': 'red', 'lw': 1.5}
            )
            ax.set_title(f"{y_col} vs {x_col}")
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)

        pairs = list(zip(x_list, y_list))
        self._save_figures_to_one_sheet(plot_one, pairs, excel_filename, sheet_name="Scatterplots")
            
    def histogram(self, x_list : list[str],n_bin : int =30,
                                excel_filename : str ="Descriptive_snapshots.xlsx"):
        """
        x_list : is a list of column names like [x1,x2,x3]
        n_bin : number of bins in histogram
        output is a histogram for each column in x_list and all histograms are saved in one sheet of excel file
        """

        if isinstance(x_list, str):
            x_list = [x_list]

   
        def plot_one(args, ax):
            x_col = args
            series = pd.to_numeric(self.df[x_col], errors='coerce').dropna()
            ax.hist(series, bins=n_bin)
            ax.set_title(f"Histogram of {x_col}")
            ax.set_xlabel(x_col)
            ax.set_ylabel("Frequency")

        pairs = x_list
        self._save_figures_to_one_sheet(plot_one, pairs, excel_filename, sheet_name="Histograms")

    def correlation_heatmap(self, columns: list[str] | str ='all',method: str ='pearson', 
                            excel_filename: str = "Descriptive_snapshots.xlsx",
                            excel_sheet_name: str = "Correlation"):
        """
        columns : list of column names to include in the correlation heatmap, or 'all' for all columns
        excel_filename : name of the Excel file to save the heatmap
        when all is selected the code only consider numeric columns for correlation 
        default method is pearson but it can be changed to spearman or kendall
        """
        if columns == 'all':
            numeric_df = self.df.select_dtypes(include='number')
            pearson_coef = numeric_df.corr(method=method)
        else:
            pearson_coef = self.df[columns].corr(method=method)
            
        # set diagonal values to zero to make it easier to see correlations between different variables
        np.fill_diagonal(pearson_coef.values, 0)

        # plot both correlation coefficients and p values in the same heatmap
        plt.figure(figsize=(10, 8))
        sns.heatmap(pearson_coef, annot=True, fmt=".2f", cmap="coolwarm", vmin=-1, vmax=1, cbar_kws={'label': 'Correlation Coefficient'})
        plt.title(f"Correlation Heatmap ({method.capitalize()})")
        plt.tight_layout()
        plt.savefig("temp_correlation_heatmap.png", bbox_inches="tight", dpi=120)
        plt.close()
        # Save the heatmap to Excel
        if os.path.exists(excel_filename):
            try:
                book = load_workbook(excel_filename)
                if excel_sheet_name in book.sheetnames:
                    book.remove(book[excel_sheet_name])
            except Exception:
                if os.path.exists(excel_filename):
                    os.remove(excel_filename)
                book = Workbook()
        else:
            book = Workbook()
        if "Sheet" in book.sheetnames and len(book.sheetnames) == 1:
            book.remove(book["Sheet"])
        ws = book.create_sheet(excel_sheet_name)
        img = XLImage("temp_correlation_heatmap.png")
        ws.add_image(img, "B2")
        book.save(excel_filename)
        book.close()
        if os.path.exists("temp_correlation_heatmap.png"):
            os.remove("temp_correlation_heatmap.png")
            
class regression_analysis:
    def __init__(self, df):
        self.df = df
        
    def best_order_finding(self, features_col : list[str],
                           target_col : str,
                           max_order : int =5,
                           criterion : str ='r2',
                           test_train_split : float =0.2,
                           method : str ='cross_val',
                           kfold : int =5, 
                           excel_filename : str ="Descriptive_snapshots.xlsx",
                           excel_sheet_name : str ="Polynomial Regression"):
        """
        features_col: list of column names to be used as features in regression
        target_col  : column name to be used as target variable in regression
        max_order   : maximum order of polynomial regression to be tested
        
        criterion   : 'r2' or 'mse' or 'all' to determine the best order of polynomial regression
                        - if 'r2' is selected the best order is the one with highest R2 score
                        - if 'mse' is selected the best order is the one with lowest MSE
                    
        method      : 'normal' or 'cross_val' to determine the method of finding best order of polynomial regression
                        - if 'normal' is selected the data is split into train and test sets and the
                              regression model is trained on the train set and evaluated on the test set
                        - if 'cross_val' is selected the data is split into k folds and the regression 
                              model is trained and evaluated on each fold and the average R2 score and MSE 
                              across all folds is used to determine the best order of polynomial regression
                            
        test_train_split : float between 0 and 1 to determine the proportion of data to be used as test set in normal method
        the function will return a dataframe with the best order of polynomial regression based on R2 score and MSE
        """
        
        df= self.df
        features=df[features_col]
        target=df[target_col]
        results = {'order': [], 'r2_score': [], 'mse': []}
        
        for order in range(1, max_order + 1):
            
            layers=[('scaling', StandardScaler()), ('poly', PolynomialFeatures(degree=order, include_bias=False)) , 
                    ('model', LinearRegression())]
            
            pipe=Pipeline(layers)

            if method == 'normal':
                X_train, X_test, y_train, y_test = train_test_split(features, target,
                                                                    test_size=test_train_split, random_state=42)
                pipe.fit(X_train, y_train)
                y_hat_test = pipe.predict(X_test)
                r2 = r2_score(y_test, y_hat_test)
                mse = mean_squared_error(y_test, y_hat_test)

            elif method == 'cross_val':
                r2_scores = cross_val_score(pipe, features, target, cv=kfold, scoring='r2')
                mse_scores = -1*cross_val_score(pipe, features, target, cv=kfold, scoring='neg_mean_squared_error')
                r2 = np.mean(r2_scores)
                mse = np.mean(mse_scores)
                print(f"Order {order} → R2: {r2:.4f}, MSE: {mse:.4f}")


            results['order'].append(order)
            results['r2_score'].append(r2)
            results['mse'].append(mse)    
            
        #plot and save the results in excel file
        plt.figure(figsize=(10, 5))
        plt.subplot(1, 2, 1)
        sns.lineplot(x=results['order'], y=results['r2_score'], marker='o')
        plt.title('R2 Score vs Polynomial Order')
        plt.xlabel('Polynomial Order')
        plt.ylabel('R2 Score')
        plt.subplot(1, 2, 2)
        sns.lineplot(x=results['order'], y=results['mse'], marker='o')
        plt.title('MSE vs Polynomial Order')
        plt.xlabel('Polynomial Order')
        plt.ylabel('MSE')
        plt.tight_layout()
        plt.savefig("temp_poly_regression_results.png", bbox_inches="tight", dpi=120)
        plt.close()
        if os.path.exists(excel_filename):
            try:
                book = load_workbook(excel_filename)
                if excel_sheet_name in book.sheetnames:
                    book.remove(book[excel_sheet_name])
            except Exception:
                if os.path.exists(excel_filename):
                    os.remove(excel_filename)
                book = Workbook()
        else:
            book = Workbook()
        if "Sheet" in book.sheetnames and len(book.sheetnames) == 1:
            book.remove(book["Sheet"])
        ws = book.create_sheet(excel_sheet_name)
        img = XLImage("temp_poly_regression_results.png")   
        ws.add_image(img, "B2")
        book.save(excel_filename)
        book.close()
        if os.path.exists("temp_poly_regression_results.png"):
            os.remove("temp_poly_regression_results.png")
        
        if criterion == 'r2':
            best_order = results['order'][np.argmax(results['r2_score'])]
        elif criterion == 'mse':
            best_order = results['order'][np.argmin(results['mse'])]
            
        print(f"Best polynomial order based on {criterion.upper()}: {best_order}")
        return best_order
    
    def regression_fitting(self, features_col : list[str], target_col : str, order : int =1):
        """
        features_col: list of column names to be used as features in regression
        target_col  : column name to be used as target variable in regression
        order       : order of polynomial regression to be fitted
        
        the function will return the fitted regression model and the R2 score and MSE of the fitted model on the whole dataset
        """
        
        df= self.df
        features=df[features_col]
        target=df[target_col]
        
        layers=[('scaling', StandardScaler()), ('poly', PolynomialFeatures(degree=order, include_bias=False)) , 
                ('model', LinearRegression())]
        
        pipe=Pipeline(layers)
        
        pipe.fit(features, target)
        y_hat = pipe.predict(features)
        r2 = r2_score(target, y_hat)
        mse = mean_squared_error(target, y_hat)
        
        print(f"Fitted polynomial regression of order {order} → R2: {r2:.4f}, MSE: {mse:.4f}")
        
        #plot distribution of data and predicted values use sns.distplot and save the plot in excel file
        fig, ax = plt.subplots(figsize=(10, 5))
        sns.kdeplot(target, label='Actual', color='blue', fill=False, ax=ax)
        sns.kdeplot(y_hat, label='Predicted', color='orange', fill=False, ax=ax)

        ax.set_title(f"Distribution of Actual vs Predicted Values (Order {order})")
        ax.set_xlabel(target_col)
        ax.set_ylabel("Density")
        ax.legend()
        plt.tight_layout()
        plt.show()
        
        
        return pipe, r2, mse
            
        
    

